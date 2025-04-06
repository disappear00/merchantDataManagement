import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import os
from datetime import datetime, timedelta
import calendar
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def convert_excel_date(date_val):
    """将Excel日期数字或中文日期转换为datetime对象"""
    try:
        if pd.isna(date_val):
            return pd.NaT

        if isinstance(date_val, (int, float)):
            return datetime(1899, 12, 30) + timedelta(days=date_val)

        date_str = str(date_val).strip().replace(" ", "")
        if "月" in date_str and "日" in date_str:
            month_day = date_str.split("月")
            month = int(month_day[0])
            day = int(month_day[1].replace("日", ""))
            year = datetime.now().year
            return datetime(year, month, day)

        return pd.to_datetime(date_val)

    except Exception as e:
        print(f"日期转换错误: {date_val}, 错误: {str(e)}")
        return pd.NaT


def format_date_as_month_day(date_val):
    """将日期格式化为'x月x日'格式"""
    if pd.isna(date_val):
        return ""
    return f"{date_val.month}月{date_val.day}日"


def calculate_tax(service_fee, salary):
    """计算税金并取负"""
    try:
        if pd.isna(service_fee) or pd.isna(salary):
            return 0
        part1 = (float(service_fee) - float(salary) * 1.0442) / (1 + 0.06) * 0.06 * 0.6725
        part2 = float(salary) * 0.0442
        return -round(part1 + part2, 2)
    except Exception as e:
        print(f"税金计算错误: {str(e)}")
        return 0


def calculate_supplement_insurance(employer_insurance):
    """计算补充险：(雇主险/2.9*1.1)并取负"""
    try:
        if pd.isna(employer_insurance):
            return 0
        return -round(float(employer_insurance) / 2.9 * 1.1, 2)
    except Exception as e:
        print(f"补充险计算错误: {str(e)}")
        return 0


def get_month_range(date_series):
    """获取日期所在月份的第一天和最后一天"""
    if date_series.empty:
        return None, None
    first_date = date_series.min()
    last_day = calendar.monthrange(first_date.year, first_date.month)[1]
    return first_date.replace(day=1), first_date.replace(day=last_day)


def apply_excel_styling(workbook, sheet_name, title):
    """应用全面的Excel样式"""
    sheet = workbook[sheet_name]

    # 设置标题行
    sheet.insert_rows(1)
    sheet.merge_cells(f'A1:{openpyxl.utils.get_column_letter(len(sheet[1]))}{1}')
    title_cell = sheet.cell(row=1, column=1, value=title)

    # 标题样式
    title_cell.font = Font(name='微软雅黑', size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 公共样式定义
    common_font = Font(name='微软雅黑', size=10)
    header_font = Font(name='微软雅黑', size=10, bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 样式应用到所有行
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
        for col_idx, cell in enumerate(row, start=1):
            # 字体
            cell.font = header_font if row_idx == 2 else common_font

            # 对齐
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 边框
            cell.border = border

    # 调整列宽
    for col_idx, col in enumerate(sheet.columns, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_idx)

        for cell in col:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass

        # 根据内容调整列宽，确保最小宽度和最大宽度
        adjusted_width = max(min((max_length + 2) * 1.2, 50), 10)
        sheet.column_dimensions[column_letter].width = adjusted_width


def process_files(text_widget):
    root = tk.Tk()
    root.withdraw()

    def append_text(message):
        text_widget.insert('end', message + '\n')
        text_widget.see('end')

    # 选择年份和月份
    current_year = datetime.now().year
    current_month = datetime.now().month
    year = simpledialog.askinteger("输入", "请输入年份:", initialvalue=current_year, minvalue=2000, maxvalue=2099)
    month = simpledialog.askinteger("输入", "请输入月份:", initialvalue=current_month, minvalue=1, maxvalue=12)

    if not year or not month:
        messagebox.showerror("错误", "年份和月份不能为空")
        return

    # 文件选择
    append_text("请选择包含地区sheet的Excel文件")
    file1_path = filedialog.askopenfilename(title="选择地区数据文件", filetypes=[("Excel文件", "*.xlsx *.xls")])
    if not file1_path:
        return

    append_text("请选择计提工资Excel文件")
    file2_path = filedialog.askopenfilename(title="选择计提工资文件", filetypes=[("Excel文件", "*.xlsx *.xls")])
    if not file2_path:
        return

    append_text("请选择配送单量Excel文件")
    file3_path = filedialog.askopenfilename(title="选择配送单量文件", filetypes=[("Excel文件", "*.xlsx *.xls")])
    if not file3_path:
        return

    append_text("请选择费用明细Excel文件")
    file4_path = filedialog.askopenfilename(title="选择费用明细文件", filetypes=[("Excel文件", "*.xlsx *.xls")])
    if not file4_path:
        return

    try:
        append_text("开始读取基础数据...")
        # 读取基础数据
        xls = pd.ExcelFile(file1_path)
        sheet_names = xls.sheet_names

        # 读取工资数据
        append_text("读取工资数据...")
        salary_df = pd.read_excel(file2_path, sheet_name=0)
        salary_df['日期'] = salary_df['日期'].apply(convert_excel_date)
        salary_df.iloc[:, 1:] = -salary_df.iloc[:, 1:]  # 工资数据全部取负

        # 读取配送单量数据
        append_text("读取配送单量数据...")
        delivery_sheets = pd.ExcelFile(file3_path).sheet_names
        if '配送单量' not in delivery_sheets:
            raise ValueError(f"配送单量文件中缺少'配送单量'sheet")
        delivery_df = pd.read_excel(file3_path, sheet_name='配送单量', header=1)
        delivery_df['日期'] = delivery_df['日期'].apply(convert_excel_date)

        # 读取摊提费用数据（第二行是列名）
        append_text("读取摊提费用数据...")
        expense_sheets = pd.ExcelFile(file4_path).sheet_names
        if '摊提费用明细' not in expense_sheets:
            raise ValueError(f"费用文件中缺少'摊提费用明细'sheet")
        amortization_df = pd.read_excel(file4_path, sheet_name='摊提费用明细', header=1)

        # 查找"日均摊销金额"行
        daily_amort_row = amortization_df[amortization_df.iloc[:, 0].str.contains('日均摊销金额', na=False)]
        if daily_amort_row.empty:
            raise ValueError("摊提费用明细中未找到'日均摊销金额'行")

        # 创建地区到数值的映射字典（取负值）
        amort_dict = {}
        for col in daily_amort_row.columns[1:]:
            region = str(col).strip()
            value = daily_amort_row[col].values[0]
            if not pd.isna(value):
                amort_dict[region] = -float(value)

        # 读取当日费用支出数据（第二行为列名）
        append_text("读取当日费用支出数据...")
        if '当日费用支出' not in expense_sheets:
            raise ValueError(f"费用文件中缺少'当日费用支出'sheet")
        daily_expense_df = pd.read_excel(file4_path, sheet_name='当日费用支出', header=1)
        daily_expense_df['日期'] = daily_expense_df['日期'].apply(convert_excel_date)

        # 准备输出文件
        output_path = os.path.splitext(file1_path)[0] + f"_{year}年{month}月_processed.xlsx"
        writer = pd.ExcelWriter(output_path, engine='openpyxl')

        # 处理每个地区sheet
        append_text("开始处理每个地区sheet...")
        for sheet in sheet_names:
            append_text(f"处理 {sheet} 地区...")
            df = pd.read_excel(file1_path, sheet_name=sheet)
            if '日期' not in df.columns:
                append_text(f"跳过 {sheet} - 缺少日期列")
                continue

            # 过滤汇总行
            df = df[~df['日期'].astype(str).str.contains('合计|本月累计', na=False)]

            new_df = pd.DataFrame()
            new_df['日期'] = df['日期'].apply(convert_excel_date)

            # 服务费回款（保持正值）
            if '合计' in df.columns:
                new_df['服务费回款'] = df['合计']
            elif '服务费回款' in df.columns:
                new_df['服务费回款'] = df['服务费回款']
            else:
                new_df['服务费回款'] = 0

            # 添加补充险（从雇主险计算）
            if '雇主险(元)' in df.columns:
                new_df['补充险'] = df['雇主险(元)'].apply(calculate_supplement_insurance)
            else:
                new_df['补充险'] = 0
                append_text(f"警告: {sheet} 无雇主险数据")

            # 合并工资数据（已经是负值）
            if sheet in salary_df.columns:
                temp_salary = salary_df[['日期', sheet]].rename(columns={sheet: '计提工资'})
                new_df = pd.merge(new_df, temp_salary, on='日期', how='left')
            else:
                new_df['计提工资'] = 0
                append_text(f"警告: {sheet} 无工资数据")

            # 合并配送单量（保持原值）
            if sheet in delivery_df.columns:
                temp_delivery = delivery_df[['日期', sheet]].rename(columns={sheet: '单量'})
                new_df = pd.merge(new_df, temp_delivery, on='日期', how='left')
            else:
                new_df['单量'] = None

            # 计算税金（函数内已取负）
            new_df['税金'] = new_df.apply(
                lambda row: calculate_tax(row['服务费回款'], abs(row['计提工资'])), axis=1)

            # 添加摊提费用（已经是负值）
            sheet_clean = sheet.strip()
            new_df['摊提费用'] = amort_dict.get(sheet_clean, 0.0)
            if sheet_clean not in amort_dict:
                append_text(f"警告: {sheet} 无摊提费用数据")

            # 添加当日费用支出（改名为本日费用）
            if sheet in daily_expense_df.columns:
                temp_expense = daily_expense_df[['日期', sheet]].rename(columns={sheet: '本日费用'})
                temp_expense['本日费用'] = temp_expense['本日费用'].fillna(0)
                new_df = pd.merge(new_df, temp_expense, on='日期', how='left')
            else:
                new_df['本日费用'] = 0
                append_text(f"警告: {sheet} 无当日费用数据")

            # 添加空列
            new_df['备注'] = ''
            new_df['当日代补'] = ''
            new_df['单均回款'] = ''

            # 计算当日利润
            new_df['当日利润'] = new_df[
                ['服务费回款', '补充险', '计提工资', '单量', '税金', '摊提费用', '本日费用']].sum(
                axis=1)

            # 计算单均回款
            new_df['单均回款'] = new_df.apply(lambda row: row['服务费回款'] / row['单量'] if row['单量'] != 0 else 0, axis=1)

            # 补全当月日期
            month_start, month_end = get_month_range(new_df['日期'])
            if month_start and month_end:
                full_dates = pd.date_range(month_start, month_end, name='日期')
                new_df = new_df.set_index('日期').reindex(full_dates).reset_index()
                new_df['日期'] = new_df['日期'].apply(format_date_as_month_day)

            # 添加合计行
            sum_row = pd.DataFrame({
                '日期': ['合计'],
                '服务费回款': [new_df['服务费回款'].sum()],
                '单量': [new_df['单量'].sum()],
                '税金': [new_df['税金'].sum()],
                '计提工资': [new_df['计提工资'].sum()],
                '摊提费用': [new_df['摊提费用'].sum()],
                '补充险': [new_df['补充险'].sum()],
                '本日费用': [new_df['本日费用'].sum()],
                '当日利润': [new_df['当日利润'].sum()],
                '备注': [''],
                '当日代补': [''],
                '单均回款': ['']
            })
            new_df = pd.concat([new_df, sum_row], ignore_index=True)

            # 调整列顺序
            column_order = ['日期', '服务费回款', '单量', '税金', '计提工资', '摊提费用',
                            '补充险', '本日费用', '当日利润', '备注', '当日代补', '单均回款']
            new_df = new_df[column_order]

            # 写入sheet
            new_df.to_excel(writer, sheet_name=sheet, index=False)

        writer.close()

        # 使用openpyxl打开并应用样式
        append_text("应用Excel样式...")
        workbook = openpyxl.load_workbook(output_path)

        for sheet_name in sheet_names:
            # 创建标题（例如：高碑店2025年3月利润明细）
            title = f"{sheet_name}{year}年{month}月利润明细"
            apply_excel_styling(workbook, sheet_name, title)

        # 保存工作簿
        workbook.save(output_path)

        append_text(f"处理完成，结果已保存至:\n{output_path}")
        messagebox.showinfo("完成", f"处理完成，结果已保存至:\n{output_path}")

    except Exception as e:
        append_text(f"处理失败:\n{str(e)}")
        messagebox.showerror("错误", f"处理失败:\n{str(e)}")


def main():
    root = tk.Tk()
    root.title("文件处理程序")

    text_widget = tk.Text(root, height=20, width=80)
    text_widget.pack()

    process_files(text_widget)

    root.mainloop()


if __name__ == "__main__":
    main()