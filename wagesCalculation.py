import os
import calendar
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import re

try:
    import openpyxl
    import xlrd
except ImportError:
    print("请先安装依赖库：pip install openpyxl xlrd")
    raise


def safe_numeric_convert(value):
    """
    安全地将各种类型的值转换为数字

    :param value: 输入的值
    :return: 转换后的数字，如果无法转换则返回0
    """
    if isinstance(value, (int, float)):
        return value

    if isinstance(value, str):
        cleaned_value = value.replace(' ', '').replace(',', '')
        match = re.search(r'[-+]?(\d+(\.\d*)?|\.\d+)', cleaned_value)
        if match:
            try:
                return float(match.group(0))
            except ValueError:
                pass

    if isinstance(value, bool):
        return 1 if value else 0

    if value is None:
        return 0

    try:
        return float(value)
    except (ValueError, TypeError):
        print(f"无法转换的值: {value}")
        return 0


def excel_column_to_index(col_str):
    """
    将Excel列名（如A、B、C）转换为列索引（0, 1, 2）

    :param col_str: Excel列名
    :return: 列索引
    """
    result = 0
    for char in col_str:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result - 1


def get_cell_value(workbook, sheet_name, cell_ref):
    """
    获取单元格的值，支持不同的Excel文件类型

    :param workbook: 工作簿对象
    :param sheet_name: 工作表名称
    :param cell_ref: 单元格引用（如 'A1'）
    :return: 单元格的值
    """
    try:
        if isinstance(workbook, openpyxl.workbook.workbook.Workbook):
            sheet = workbook[sheet_name]
            return safe_numeric_convert(sheet[cell_ref].value)
        elif isinstance(workbook, xlrd.book.Book):
            sheet = workbook.sheet_by_name(sheet_name)
            col_str = ''.join(filter(str.isalpha, cell_ref))
            row_str = ''.join(filter(str.isdigit, cell_ref))
            col_index = excel_column_to_index(col_str)
            row_index = int(row_str) - 1
            return safe_numeric_convert(sheet.cell_value(row_index, col_index))
        else:
            raise ValueError("不支持的工作簿类型")
    except Exception as e:
        print(f"获取单元格值时出错: {cell_ref}, {e}")
        return 0


def calculate_area_salary(workbook, area_formula):
    """
    根据给定的Excel工作簿和地区计算公式计算工资

    :param workbook: Excel工作簿对象
    :param area_formula: 地区工资计算公式
    :return: 计算的工资值
    """
    try:
        parts = area_formula.split('+')
        salary_parts = []
        for part in parts:
            part = part.strip()
            sheet_name, cell = part.split('!')
            value = get_cell_value(workbook, sheet_name, cell)
            salary_parts.append(value)
        return round(sum(salary_parts), 2)
    except Exception as e:
        print(f"计算地区工资时出错: {e}")
        return 0.00


def generate_salary_summary(input_file, output_file, month, end_day, text_widget):
    """
    生成工资计提汇总表

    :param input_file: 输入的Excel文件路径
    :param output_file: 输出的Excel文件路径
    :param month: 计算的月份
    :param end_day: 截止日期
    :param text_widget: 用于显示信息的Text组件
    """
    wb = load_workbook(input_file)
    areas = {
        '高碑店': '运营中心!W6+摊销人员!E6+后线及站长!M6+业务侧薪资汇总!B3+配送总表!B3',
        '白沟': '运营中心!X6+摊销人员!F6+后线及站长!AC6+业务侧薪资汇总!C3+配送总表!C3',
        '新城': '新城工资!F2+配送总表!D3',
        '霸州': '运营中心!Y6+摊销人员!M6+后线及站长!O46+业务侧薪资汇总!E3+配送总表!E3',
        '胜芳': '运营中心!Z6+摊销人员!N6+后线及站长!X46+业务侧薪资汇总!F3+配送总表!F3',
        '霸州乡镇': '配送总表!G3',
        '邢台': '运营中心!AA6+后线及站长!AD86+业务侧薪资汇总!G3+配送总表!H3',
        '下花园': '运营中心!AC6+摊销人员!U6+后线及站长!G125+业务侧薪资汇总!H3+配送总表!I3',
        '万全': '运营中心!AD6+摊销人员!T6+后线及站长!O125+业务侧薪资汇总!I3+配送总表!J3'
    }
    new_wb = openpyxl.Workbook()
    sheet = new_wb.active
    sheet.title = '工资计提汇总'
    headers = ['日期'] + list(areas.keys()) + ['合计']
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)
    _, max_days = calendar.monthrange(datetime.now().year, month)
    for day in range(1, max_days + 1):
        row = day + 1
        sheet.cell(row=row, column=1, value=f"{month}月{day}日")
        if day <= end_day:
            row_total = 0
            for col, (area, formula) in enumerate(areas.items(), 2):
                adjusted_formula = []
                for part in formula.split('+'):
                    part = part.strip()
                    sheet_name, cell_ref = part.split('!')
                    col_str = ''.join(filter(str.isalpha, cell_ref))
                    row_num = ''.join(filter(str.isdigit, cell_ref))
                    if row_num:
                        new_row_num = str(int(row_num) + (day - 1))
                        adjusted_cell = f"{col_str}{new_row_num}"
                        adjusted_formula.append(f"{sheet_name}!{adjusted_cell}")
                    else:
                        adjusted_formula.append(part)
                adjusted_formula = '+'.join(adjusted_formula)
                salary = calculate_area_salary(wb, adjusted_formula)
                sheet.cell(row=row, column=col, value=round(salary, 2))
                row_total += salary
            sheet.cell(row=row, column=len(headers), value=round(row_total, 2))
        else:
            for col in range(2, len(headers) + 1):
                sheet.cell(row=row, column=col, value="")
    new_wb.save(output_file)
    text_widget.insert(tk.END, f"工资计提汇总表已生成: {output_file}\n")
    text_widget.see(tk.END)


def select_input_file(text_widget):
    """
    通过文件对话框选择输入文件
    """
    input_file = filedialog.askopenfilename(
        title="选择Excel输入文件(总商薪资摊销2025.3终(1))",
        filetypes=[
            ("Excel文件", "*.xlsx *.xls *.xlsm *.xlsb *.xltx *.xltm"),
            ("所有文件", "*.*")
        ]
    )
    if not input_file:
        return
    month = simpledialog.askinteger(
        "选择月份",
        "请输入要计算的月份(1-12):",
        minvalue=1,
        maxvalue=12,
        initialvalue=datetime.now().month
    )
    if not month:
        return
    _, max_days = calendar.monthrange(datetime.now().year, month)
    end_day = simpledialog.askinteger(
        "选择截止日期",
        f"请输入要生成数据的截止日期(1-{max_days}):",
        minvalue=1,
        maxvalue=max_days,
        initialvalue=max_days
    )
    if not end_day:
        return
    output_file = filedialog.asksaveasfilename(
        title="保存工资汇总表",
        defaultextension=".xlsx",
        filetypes=[("Excel文件", "*.xlsx")]
    )
    if not output_file:
        return
    try:
        generate_salary_summary(input_file, output_file, month, end_day, text_widget)
    except Exception as e:
        messagebox.showerror("错误", str(e))


def load_workbook(file_path):
    """
    根据文件扩展名加载不同类型的Excel文件

    :param file_path: Excel文件路径
    :return: 工作簿对象
    """
    file_extension = os.path.splitext(file_path)[1].lower()
    if file_extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        return openpyxl.load_workbook(file_path, data_only=True)
    elif file_extension in ['.xls', '.xlsb']:
        return xlrd.open_workbook(file_path)
    else:
        raise ValueError(f"不支持的文件类型: {file_extension}")


def main():
    root = tk.Tk()
    root.title("工资计提汇总表生成程序")
    text_widget = tk.Text(root, height=20, width=80)
    text_widget.pack()

    def append_text(message):
        text_widget.insert(tk.END, message + "\n")
        text_widget.see(tk.END)

    append_text("请选择Excel输入文件")
    select_input_file(text_widget)
    root.mainloop()


if __name__ == '__main__':
    main()